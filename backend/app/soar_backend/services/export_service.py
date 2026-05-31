"""
export_service.py — Multi-format report export engine.

Supports:
  - PDF (FPDF2 — professional report with cover page, sections, tables, charts)
  - CSV (multi-section, table-format)
  - JSON (full structured)
  - XLSX (multi-sheet Excel workbook)
"""

import json
import logging
from pathlib import Path
from datetime import datetime, timezone

logger = logging.getLogger(__name__)

EXPORT_DIR = Path("/tmp/cybernest_reports")


class ExportResult:
    path: str
    mime_type: str
    size_bytes: int
    filename: str

    def __init__(self, path: str, mime_type: str, size_bytes: int, filename: str):
        self.path = path
        self.mime_type = mime_type
        self.size_bytes = size_bytes
        self.filename = filename


class ExportService:
    """Multi-format export engine for structured report data."""

    def __init__(self):
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    async def export(self, report: dict, fmt: str) -> ExportResult:
        exporters = {
            "pdf": self._export_pdf,
            "csv": self._export_csv,
            "json": self._export_json,
            "xlsx": self._export_xlsx,
        }
        exporter = exporters.get(fmt)
        if not exporter:
            raise ValueError(f"Unsupported format: {fmt}")
        return await exporter(report)

    def _flatten_sections(self, report: dict, parent_key: str = "") -> list[tuple[str, str, str, str]]:
        """Flatten nested report dict into Section, SubSection, Metric, Value rows."""
        rows = []
        stack = [(parent_key, report)]

        while stack:
            key, value = stack.pop()
            if isinstance(value, dict):
                for k, v in value.items():
                    if isinstance(v, (dict, list)):
                        stack.append((f"{key}/{k}" if key else k, v))
                    else:
                        rows.append((key or "root", k, str(v), ""))
            elif isinstance(value, list):
                for i, item in enumerate(value):
                    if isinstance(item, dict):
                        for k, v in item.items():
                            if not isinstance(v, (dict, list)):
                                rows.append((f"{key}[{i}]", k, str(v), ""))
                    else:
                        rows.append((key, str(i), str(item), ""))
            else:
                rows.append((key, "", str(value), ""))

        return rows

    async def _export_csv(self, report: dict) -> ExportResult:
        """Multi-section CSV with Section, SubSection, Metric, Value columns."""
        report_type = report.get("report_type", "report")
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"cybernest-{report_type}-{timestamp}.csv"
        filepath = str(EXPORT_DIR / filename)

        def esc(v):
            return f'"{str(v or "").replace('"', '""')}"'

        rows = ["Section,Sub-Section,Metric,Value"]
        for section, sub, metric, val in self._flatten_sections(report):
            rows.append(f"{esc(section)},{esc(sub)},{esc(metric)},{esc(val)}")

        content = "\n".join(rows)
        Path(filepath).write_text(content)
        return ExportResult(filepath, "text/csv", len(content), filename)

    async def _export_json(self, report: dict) -> ExportResult:
        """Full structured JSON export."""
        report_type = report.get("report_type", "report")
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"cybernest-{report_type}-{timestamp}.json"
        filepath = str(EXPORT_DIR / filename)

        content = json.dumps(report, indent=2, default=str)
        Path(filepath).write_text(content)
        return ExportResult(filepath, "application/json", len(content), filename)

    async def _export_pdf(self, report: dict) -> ExportResult:
        """Professional PDF with cover page, sections, and data tables.

        Uses FPDF2 for server-side generation (no browser dependency).
        """
        try:
            from fpdf import FPDF
        except ImportError:
            logger.warning("FPDF2 not installed. Install with: pip install fpdf2")
            # Fall back to a simple text-based PDF
            return await self._export_txt_pdf(report)

        report_type = report.get("report_type", "report").replace("_", " ").title()
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"cybernest-{report.get('report_type', 'report')}-{timestamp}.pdf"
        filepath = str(EXPORT_DIR / filename)

        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.add_page()

        # Cover page
        pdf.set_fill_color(10, 14, 26)
        pdf.rect(0, 0, 210, 297, "F")
        pdf.set_text_color(0, 200, 255)
        pdf.set_font("Helvetica", "B", 28)
        pdf.cell(0, 40, "CYBERNEST SOAR", align="C")
        pdf.ln(50)
        pdf.set_font("Helvetica", "", 20)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(0, 15, report_type, align="C")
        pdf.ln(20)
        pdf.set_font("Helvetica", "", 12)
        pdf.set_text_color(180, 180, 180)
        period = report.get("period", {})
        pdf.cell(0, 10, f"Period: {period.get('start', 'N/A')[:10]} — {period.get('end', 'N/A')[:10]}", align="C")
        pdf.ln(10)
        pdf.cell(0, 10, f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", align="C")
        pdf.ln(20)
        pdf.set_text_color(100, 100, 100)
        pdf.set_font("Helvetica", "I", 10)
        pdf.cell(0, 10, "CONFIDENTIAL — For Authorized Personnel Only", align="C")

        # Content sections
        for section_key, section_data in report.items():
            if section_key in ("report_id", "report_type", "generated_at", "period"):
                continue
            if not isinstance(section_data, (dict, list)):
                continue

            pdf.add_page()
            pdf.set_text_color(0, 200, 255)
            pdf.set_font("Helvetica", "B", 16)
            pdf.cell(0, 12, section_key.replace("_", " ").title())
            pdf.ln(15)

            if isinstance(section_data, dict):
                for k, v in section_data.items():
                    if isinstance(v, (dict, list)):
                        continue
                    pdf.set_text_color(200, 200, 200)
                    pdf.set_font("Helvetica", "", 10)
                    pdf.cell(80, 7, k.replace("_", " ").title())
                    pdf.set_font("Helvetica", "B", 10)
                    pdf.set_text_color(255, 255, 255)
                    pdf.cell(0, 7, str(v)[:60])
                    pdf.ln(7)

        pdf.output(filepath)
        size = Path(filepath).stat().st_size
        return ExportResult(filepath, "application/pdf", size, filename)

    async def _export_txt_pdf(self, report: dict) -> ExportResult:
        """Fallback plain-text PDF when FPDF2 is not available."""
        report_type = report.get("report_type", "report")
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"cybernest-{report_type}-{timestamp}.pdf"
        filepath = str(EXPORT_DIR / filename)

        lines = [
            "=" * 80,
            "CYBERNEST SOAR REPORT",
            "=" * 80,
            f"Type: {report_type}",
            f"Generated: {datetime.now(timezone.utc).isoformat()}",
            "=" * 80,
            "",
        ]
        for section_key, section_data in report.items():
            if section_key in ("report_id", "report_type", "generated_at", "period"):
                continue
            lines.append(f"\n{'=' * 40}")
            lines.append(f"  {section_key.replace('_', ' ').title()}")
            lines.append(f"{'=' * 40}")
            if isinstance(section_data, dict):
                for k, v in section_data.items():
                    if not isinstance(v, (dict, list)):
                        lines.append(f"  {k.replace('_', ' ').title()}: {v}")

        content = "\n".join(lines)
        Path(filepath).write_text(content)
        return ExportResult(filepath, "application/pdf", len(content), filename)

    async def _export_xlsx(self, report: dict) -> ExportResult:
        """Multi-sheet Excel workbook.

        Each top-level section becomes a sheet.
        Requires openpyxl.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            logger.warning("openpyxl not installed. Install with: pip install openpyxl")
            raise ValueError("openpyxl required for XLSX export")

        report_type = report.get("report_type", "report")
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"cybernest-{report_type}-{timestamp}.xlsx"
        filepath = str(EXPORT_DIR / filename)

        wb = Workbook()
        wb.remove(wb.active)

        for section_key, section_data in report.items():
            if section_key in ("report_id", "report_type", "generated_at", "period"):
                continue
            if not isinstance(section_data, (dict, list)):
                continue

            sheet_name = section_key[:31].replace("_", " ").title()
            ws = wb.create_sheet(title=sheet_name)

            if isinstance(section_data, dict):
                row_num = 1
                for k, v in section_data.items():
                    if isinstance(v, (dict, list)):
                        continue
                    ws.cell(row=row_num, column=1, value=k.replace("_", " ").title())
                    cell = ws.cell(row=row_num, column=2, value=str(v))
                    cell.font = Font(bold=True)
                    row_num += 1

        wb.save(filepath)
        size = Path(filepath).stat().st_size
        return ExportResult(filepath,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           size, filename)
